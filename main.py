#!/usr/bin/python3
import os
import time
import random
from random import shuffle
from time import gmtime, strftime

from openpyxl import Workbook
import openpyxl

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

HOMEPAGE = 'https://translate.google.com/'
DRIVER_PATH = './chrome_driver/chromedriver.exe'

#inputs
TURNS = 20000

XLSX = './output/output ' + str(strftime("%Y-%m-%d %H-%M-%S", gmtime())) + '.xlsx'

def load_driver():
    if os.path.isfile(DRIVER_PATH):
        print('loading driver...')
    else:
        print("can't locate driver")
    driver = webdriver.Chrome(executable_path=DRIVER_PATH)
    driver.get(HOMEPAGE)

    print(driver.current_url)

    if not os.path.exists('./output'):
        os.makedirs('./output')

    #create workbook
    wb = Workbook()
    wb.save(XLSX)

    post(driver)


def key_gen(size):

    key = ''

    for i in range(1):

        chrlist = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                   'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                   'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm',
                   'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']

        symbols = ['!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '-', '=', '+',
                   '.', ':', ';', '[', ']', '"', "'", "/", '¥', '§', '£','¢', '`',
                   '~']

        numb = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']

        japanese = ['ろ', 'ぬ', 'ふ', 'あ', 'う', 'え', 'お', 'や', 'ゆ', 'よ', 'わ',
                    'ほ', 'へ', 'た', 'て', 'い', 'す', 'か', 'ん', 'な', 'に', 'ら', 'ぜ',
                    '゜', 'む', 'ち', 'と', 'は', 'し', 'き', 'く', 'ま', 'の', 'り', 'れ',
                    'け', 'む', 'つ', 'さ', 'そ', 'ひ', 'こ', 'み', 'も', 'ね', 'る', 'め']

        chinesee = ['潍', '敲', '挠', '慨', '慲', '瑣', '牥', 'ഠ', '上', '呏', '㩅', '字',
                    '朠', '瑥', '楴', '杮', '洠', '牯', '档', '牡', '捡', '整', '这', '中',
                    '獲', '眠', '汩', '慴', '敫', '猠', '浯', '楴', '敭', '月', '再', '张',
                    '安', '吧', '八', '爸', '百', '北', '不', '大', '岛', '的', '弟', '地',
                    '东', '都', '对', '多', '儿', '二', '方', '港', '哥', '个', '关', '贵',
                    '国', '过', '海', '好', '很', '会', '家', '见', '叫', '姐', '京', '九',
                    '可', '老', '李', '零', '六', '吗', '妈', '么', '没', '美', '妹', '们',
                    '名', '明', '哪', '那', '南', '你', '您', '朋', '七', '起', '千', '去',
                    '人', '认', '日', '三', '上', '谁', '什', '生', '师', '十', '识', '是',
                    '四', '他', '她', '台', '天', '湾', '万', '王', '我', '五', '西', '息',
                    '系', '先', '香', '想', '小', '谢', '姓', '休', '学', '也', '一', '亿',
                    '英', '友']

        hebrew = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'י', 'ך', 'כ', 'ל', 'ם',
                  'מ', 'ן', 'נ', 'ס', 'ע', 'ף', 'פ', 'ץ', 'צ', 'ק', 'ר', 'ש', 'ת']

        arabic = ['ذ', '١', '٢', '٣', '٤', '٥', '٦', '٧', '٨', '٩', '٠', 'ض', 'ص', 'ث',
                  'ق', 'ف', 'غ', 'ع', 'ه', 'خ', 'ح', 'ج', 'د', 'ش', 'س', 'ي', 'ب', 'ل',
                  'ا', 'ت', 'ن', 'م', 'ك', 'ط', 'ئ', 'ء', 'ؤ', 'ر', 'ل', 'ى', 'ة', 'و',
                  'ز', 'ظ', 'ل', 'أ-،إ;ًَ', 'أ', 'ا']

        hindi = ['ौ', 'ै', 'ा', 'ी', 'ू', 'ब', 'ह', 'ग', 'द', 'ज', 'ड', 'ॉ', 'ो', 'े',
                 '्', 'ि', 'ु', 'प', 'र', 'क', 'त', 'च', 'ट', 'य', '़', 'स', 'ल', 'व',
                'न', 'म', 'ं', 'ॉ']

        russian = ['й', 'ц', 'у', 'к', 'е', 'н', 'г', 'ш', 'щ', 'з', 'х', 'ъ', 'ф', 'ы',
                   'ы', 'в', 'а', 'п', 'р', 'о', 'л', 'д', 'ж', 'э', 'я', 'ч', 'с', 'м',
                   'и', 'т', 'ь', 'б', 'ю', 'ё',]

        the_list = []

        the_list.append(hebrew)
        shuffle(hebrew)
        the_list.append(arabic)
        shuffle(arabic)
        the_list.append(chinesee)
        shuffle(chinesee)
        the_list.append(japanese)
        shuffle(japanese)
        the_list.append(hindi)
        shuffle(hindi)
        the_list.append(numb)
        shuffle(numb)
        the_list.append(symbols)
        shuffle(symbols)
        the_list.append(chrlist)
        shuffle(chrlist)
        the_list.append(russian)
        shuffle(russian)

        #number of lists 0-8 #random 2-4 chars to choose from
        em = the_list[int(random.randint(0, 8))][:int(random.randint(2, 4))]

        print(em)

        for i in range(size):

            me = the_list[int(random.randint(0, 8))][:int(random.randint(2, 4))]

            #to choose from
            #choice(me) to choose from same list
            #choice(em) mixed
            key += ''.join(random.choice(em) for _ in range(random.randint(2, 3)))  # chars per word 2-3
            key += ' '  # adds spcae

    return key


def post(driver):

    counter = 0

    while counter < TURNS:
        
        counter += 1
        print(counter)
        source = driver.find_element(By.XPATH, '//*[@id="source"]')
        key = (key_gen(int(random.randint(15, 45)))) #char counter 15-45
        source.send_keys(key)
        time.sleep(0.5)
        global_var1 = key

        try:
            try:
                time.sleep(1.5)
                link = driver.find_element(By.XPATH, '/ html / body')
                link.click()
                result = driver.find_element(By.XPATH,
                                             "./html/body/div/div/form/div/div/div/div/div/div/div/div/div/div/span[@id='result_box']/span")
                re = result.text
                global_var2 = (str(re))
                wb = openpyxl.load_workbook(XLSX)
                # grab the active worksheet
                ws = wb.active
                if global_var2.strip(' ') == global_var1.strip(' '):
                    ws['C' + str(counter)].value = 'same_string'

                print('Output: ' + global_var2)
                print('Input: ' + global_var1)

                # Data can be assigned directly to cells
                ws['A' + str(counter)].value = str(global_var1)
                ws['B' + str(counter)].value = str(global_var2)

                # lang detection
                detection = driver.find_element(By.XPATH, './/*[@id="gt-sl-sugg"]/div/div[5]')
                de = detection.text
                detected = (str(de))
                print('Language: ' + detected)
                ws['D' + str(counter)].value = str(detected)

                # Save the file
                wb.save(XLSX)

            except NoSuchElementException:

                print('ElementException')
                time.sleep(8)
                result = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,
                                                    "./html/body/div/div/form/div/div/div/div/div/div/div/div/div/div/span[@id='result_box']/span"))
                )
                link = driver.find_element(By.XPATH, '/ html / body')
                link.click()
                re = result.text
                global_var2 = (str(re))
                wb = openpyxl.load_workbook(XLSX)

                # grab the active worksheet
                ws = wb.active

                if global_var2.strip(' ') == global_var1.strip(' '):
                    print('same_string')
                    ws['C' + str(counter)].value = 'same_string'

                print(global_var2)
                print(global_var1)

                # Data can be assigned directly to cells
                ws['A' + str(counter)].value = str(global_var1)
                ws['B' + str(counter)].value = str(global_var2)
                ws['E' + str(counter)].value = 'NoSuchElementException'

                # Save the file
                wb.save(XLSX)

        except TimeoutException:

            print('TimeoutException')

            pass

        driver.get(HOMEPAGE)

load_driver()
